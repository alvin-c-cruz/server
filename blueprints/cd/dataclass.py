from flask import flash, current_app
from datetime import date
from dataclasses import dataclass
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .. DB import get_db
from ... packages import Voucher

@dataclass
class CD(Voucher):
    cd_num: str = ""
    record_date: str = str(date.today())
    vendor_id: int = 0
    check_number: str = ""
    description: str = ""

    @property
    def vendor_name(self):
    	return self.db.execute('SELECT name FROM tbl_vendor WHERE id=?', (self.vendor_id, )).fetchone()[0]	
		

    def all(self, **filter):
    	if filter:
    		clause = [f'{key}=?' for key in filter]
    		sql = f"""
    			SELECT 
    				tbl_cd.id,
                    tbl_cd.cd_num,
    				tbl_cd.record_date,
    				tbl_vendor.name as vendor_name,
                    tbl_cd.check_number,
    				tbl_cd.description
    			FROM tbl_cd
    			INNER JOIN tbl_vendor ON tbl_vendor.id = tbl_cd.vendor_id
    			WHERE {", ".join(clause)}
    		"""
    		return self.db.execute(sql, tuple(filter.values)).fetchall()
    	else:
    		sql = f"""
    			SELECT 
    				tbl_cd.id,
    				tbl_cd.cd_num,
                    tbl_cd.record_date,
                    tbl_cd.check_number,
    				tbl_vendor.name as vendor_name,
    				tbl_cd.description
    			FROM tbl_cd
    			INNER JOIN tbl_vendor ON tbl_vendor.id = tbl_cd.vendor_id
    		"""
    		return self.db.execute(sql).fetchall()


    def range(self, date_from, date_to):
        if True:
            sql = f"""
                SELECT 
                    tbl_cd.id,
                    tbl_cd.cd_num,
                    tbl_cd.record_date,
                    tbl_vendor.name as vendor_name,
                    tbl_cd.check_number,
                    tbl_cd.description
                FROM tbl_cd
                INNER JOIN tbl_vendor ON tbl_vendor.id = tbl_cd.vendor_id
                WHERE tbl_cd.record_date>=? AND tbl_cd.record_date<=?
            """
            return self.db.execute(sql, (date_from, date_to)).fetchall()


    def is_validated(self):
        if self.id:
            if self.db.execute('SELECT COUNT(*) FROM tbl_cd WHERE cd_num=? AND id!=?;', (self.cd_num, self.id)).fetchone()[0]:
                flash("CD Number is already in use.")
                return False
            if self.db.execute('SELECT COUNT(*) FROM tbl_cd WHERE check_number=? AND id!=?;', (self.check_number, self.id)).fetchone()[0]:
                flash("Check Number is already in use.")
                return False
        else:
            if self.db.execute('SELECT COUNT(*) FROM tbl_cd WHERE cd_num=?;', (self.cd_num, )).fetchone()[0]:
                flash("CD Number is already in use.")
                return False
            if self.db.execute('SELECT COUNT(*) FROM tbl_cd WHERE check_number=?;', (self.check_number, )).fetchone()[0]:
                flash("Check Number is already in use.")
                return False
                
        return True


@dataclass
class Create_File:
    date_from: str
    date_to: str

    def __post_init__(self):
        self.filename = os.path.join(current_app.instance_path, "downloads", f"{self.date_from} to {self.date_to} CD.xlsx")
        db = get_db()
        self.cds = CD(db).range(self.date_from, self.date_to)

        #  Create initial dataframe
        sql = f"""SELECT 
                    cd.id as id,
                    cd.record_date as DATE, 
                    cd.cd_num AS "CD No.", 
                    v.name as NAME, 
                    cd.check_number as "CHECK No.",
                    cd.description as DESCRIPTION
                
                FROM tbl_cd as cd
                
                INNER JOIN tbl_vendor as v ON cd.vendor_id = v.id

                WHERE cd.record_date>="{self.date_from}" AND cd.record_date<="{self.date_to}"
            ;
            """
        df_cd = pd.read_sql_query(sql, db)
        df_cd = df_cd.set_index('id')

        #  Collect account titles and add to main dataframe
        sql = f"""SELECT 
                    acct.name as account_title
                
                FROM tbl_cd_entry as entry
                
                INNER JOIN tbl_cd as cd ON entry.cd_id = cd.id
                INNER JOIN tbl_account as acct on acct.id = entry.account_id

                WHERE cd.record_date>="{self.date_from}" AND cd.record_date<="{self.date_to}"

                GROUP BY acct.name

                ORDER BY acct.account_number
            ;
            """
        df_accounts = pd.read_sql_query(sql, db)
        list_accounts = []
        for key, row in df_accounts.iterrows():
            account_title = row.account_title
            list_accounts.append(account_title.upper())

        for col_name in list_accounts: df_cd[col_name] = ""

        #  Gather entries and record to proper row and column
        sql = f"""SELECT 
                    cd.id, 
                    acct.name as account_title,
                    (entry.debit-entry.credit) as amount 
                
                FROM tbl_cd_entry as entry
                
                INNER JOIN tbl_cd as cd ON entry.cd_id = cd.id
                INNER JOIN tbl_account as acct on acct.id = entry.account_id

                WHERE cd.record_date>="{self.date_from}" AND cd.record_date<="{self.date_to}"
            ;
            """
        df_entry = pd.read_sql_query(sql, db)
        df_entry = df_entry.set_index('id')

        for key, row in df_entry.iterrows():
            account_title, amount = row
            df_cd.loc[key][account_title.upper()] = amount

        self.df_cd = df_cd

        self.create()


    def create(self):
        wb = Workbook()

        self.sheet_transactions(wb)

        wb.save(self.filename)
        wb.close()


    def sheet_transactions(self, wb):
        sheet_name = "CDJ"
        wb["Sheet"].title = sheet_name
        ws = wb[sheet_name]

        #  Headers
        row_num = 1
        column_names = list(self.df_cd.columns)
        
        col_num = 1
        for col_name in column_names:
            cell = ws[f'{get_column_letter(col_num)}{row_num}']
            cell.value = col_name
            col_num += 1



        #  Details
        row_num = 2
        for key, cd in self.df_cd.iterrows():
            col_num = 1
            for x in cd:
                cell = ws[f'{get_column_letter(col_num)}{row_num}']
                cell.value = x
                col_num += 1
            
            row_num += 1