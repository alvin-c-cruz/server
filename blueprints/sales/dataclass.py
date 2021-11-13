from flask import flash, current_app, session, g
from datetime import date
from dataclasses import dataclass
import os
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side

from .. DB import get_db
from ... packages import Voucher

@dataclass
class CD(Voucher):
    sv_num: str = ""
    record_date: str = str(date.today())
    customer_id: int = 0
    invoice_num: str = ""
    description: str = ""

    @property
    def customer_name(self):
    	return self.db.execute('SELECT name FROM tbl_customer WHERE id=?', (self.customer_id, )).fetchone()[0]	
		

    def all(self, **filter):
    	if filter:
    		clause = [f'{key}=?' for key in filter]
    		sql = f"""
    			SELECT 
    				tbl_sales.id,
                    tbl_sales.sv_num,
    				tbl_sales.record_date,
    				tbl_customer.name as customer_name,
                    tbl_sales.invoice_num,
    				tbl_sales.description
    			FROM tbl_sales
    			INNER JOIN tbl_customer ON tbl_customer.id = tbl_sales.customer_id
    			WHERE {", ".join(clause)}
    		"""
    		return self.db.execute(sql, tuple(filter.values)).fetchall()
    	else:
    		sql = f"""
    			SELECT 
    				tbl_cd.id,
    				tbl_cd.cd_num,
                    tbl_cd.record_date,
                    tbl_cd.check_number,
    				tbl_vendor.name as vendor_name,
    				tbl_cd.description
    			FROM tbl_cd
    			INNER JOIN tbl_vendor ON tbl_vendor.id = tbl_cd.vendor_id
    		"""
    		return self.db.execute(sql).fetchall()


    def range(self, date_from, date_to):
        if True:
            sql = f"""
                SELECT 
                    tbl_cd.id,
                    tbl_cd.cd_num,
                    tbl_cd.record_date,
                    tbl_vendor.name as vendor_name,
                    tbl_cd.check_number,
                    tbl_cd.description
                FROM tbl_cd
                INNER JOIN tbl_vendor ON tbl_vendor.id = tbl_cd.vendor_id
                WHERE tbl_cd.record_date>=? AND tbl_cd.record_date<=?
            """
            return self.db.execute(sql, (date_from, date_to)).fetchall()


    def is_validated(self):
        if self.id:
            if self.db.execute('SELECT COUNT(*) FROM tbl_cd WHERE cd_num=? AND id!=?;', (self.cd_num, self.id)).fetchone()[0]:
                flash("CD Number is already in use.")
                return False
            if self.db.execute('SELECT COUNT(*) FROM tbl_cd WHERE check_number=? AND id!=?;', (self.check_number, self.id)).fetchone()[0]:
                flash("Check Number is already in use.")
                return False
        else:
            if self.db.execute('SELECT COUNT(*) FROM tbl_cd WHERE cd_num=?;', (self.cd_num, )).fetchone()[0]:
                flash("CD Number is already in use.")
                return False
            if self.db.execute('SELECT COUNT(*) FROM tbl_cd WHERE check_number=?;', (self.check_number, )).fetchone()[0]:
                flash("Check Number is already in use.")
                return False
                
        return True


def get_cd(date_from, date_to):
    def short_date(_date):
        if type(_date) == str:
            _date = date(int(_date[:4]), int(_date[5:7]), int(_date[-2:]))

        return _date.strftime("%d-%b-%Y")

    #  Open model        
    db = get_db()

    #  Create initial dataframe
    sql = f"""SELECT 
                cd.id as id,
                cd.record_date as DATE, 
                cd.cd_num AS "CD No.", 
                v.name as NAME, 
                cd.check_number as "CHECK No.",
                cd.description as DESCRIPTION
            
            FROM tbl_cd as cd
            
            INNER JOIN tbl_vendor as v ON cd.vendor_id = v.id

            WHERE cd.record_date>="{date_from}" AND cd.record_date<="{date_to}"
        ;
        """
    df_cd = pd.read_sql_query(sql, db)
    df_cd = df_cd.set_index('id')

    for key, row in df_cd.iterrows():
        row.DATE = short_date(row.DATE)


    #  Collect account titles and add to main dataframe
    sql = f"""SELECT 
                acct.name as account_title
            
            FROM tbl_cd_entry as entry
            
            INNER JOIN tbl_cd as cd ON entry.cd_id = cd.id
            INNER JOIN tbl_account as acct on acct.id = entry.account_id

            WHERE cd.record_date>="{date_from}" AND cd.record_date<="{date_to}"

            GROUP BY acct.name

            ORDER BY acct.account_number
        ;
        """
    df_accounts = pd.read_sql_query(sql, db)
    list_accounts = []
    for key, row in df_accounts.iterrows():
        account_title = row.account_title
        list_accounts.append(account_title.upper())

    for col_name in list_accounts: df_cd[col_name] = ""

    #  Gather entries and record to proper row and column
    sql = f"""SELECT 
                cd.id, 
                acct.name as account_title,
                (entry.debit-entry.credit) as amount 
            
            FROM tbl_cd_entry as entry
            
            INNER JOIN tbl_cd as cd ON entry.cd_id = cd.id
            INNER JOIN tbl_account as acct on acct.id = entry.account_id

            WHERE cd.record_date>="{date_from}" AND cd.record_date<="{date_to}"
        ;
        """
    df_entry = pd.read_sql_query(sql, db)
    df_entry = df_entry.set_index('id')

    for key, row in df_entry.iterrows():
        account_title, amount = row
        df_cd.loc[key][account_title.upper()] = amount

    return df_cd


@dataclass
class Create_File:
    date_from: str
    date_to: str

    def __post_init__(self):
        #  Excel formats
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.double_rule = Border(
            bottom=Side(style='double')
        )

        #  Downloaded filename
        self.filename = os.path.join(current_app.instance_path, "downloads", f"{self.date_from} to {self.date_to} Cash Disbursement Journal.xlsx")
        

        self.df_cd = get_cd(self.date_from, self.date_to)

        self.create()


    def create(self):
        wb = Workbook()

        self.sheet_transactions(wb)

        wb.save(self.filename)
        wb.close()


    def sheet_transactions(self, wb):
        sheet_name = "CDJ"
        wb["Sheet"].title = sheet_name
        ws = wb[sheet_name]

        #  Report head
        row_num = 1
        cell = ws[f'A{row_num}']
        cell.value = session.get('company_name')
        cell.font = Font(size=14, bold=True)

        row_num += 1
        cell = ws[f'A{row_num}']
        cell.value = "Cash Disbursement Journal"
        cell.font = Font(size=12, bold=True)

        row_num += 1
        cell = ws[f'A{row_num}']
        cell.value = f"From {self.long_date(self.date_from)} to {self.long_date(self.date_to)}"
        cell.font = Font(size=12, bold=True)

        #  Headers
        row_num += 2
        column_names = list(self.df_cd.columns)
        
        width = {
            "A": 11.11,
            "B": 9.11,
            "C": 30.0,
            "D": 9.11,
            "E": 30.0,
        }
            

        col_num = 1
        for col_name in column_names:
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = width[col_letter] if col_letter in width else 14
            cell = ws[f'{col_letter}{row_num}']
            cell.value = col_name
            cell.font = Font(size=10, bold=True)
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center') if col_num < 6 else Alignment(horizontal='center', wrap_text=True)
            col_num += 1



        #  Details
        row_num += 1
        start_row = row_num
        amount_columns = []
        for key, cd in self.df_cd.iterrows():
            col_num = 1
            
            for x in cd:
                col_letter = get_column_letter(col_num)
                cell = ws[f'{col_letter}{row_num}']
                cell.value = x
                cell.border = self.thin_border
                if col_letter in ("A", "B", "D"): cell.alignment = Alignment(horizontal='center')
                if col_letter in ("B", "D"): cell.number_format = "@"

                if col_num not in (1, 2, 3, 4, 5): 
                    amount_columns.append(col_letter)
                    cell.number_format = "#,##0.00_);(#,##0.00)"
                col_num += 1
            
            row_num += 1
        
        end_row = row_num

        #  Total
        row_num += 1
        cell = ws[f'A{row_num}']
        cell.value = "TOTAL"
        
        for col in ("A", "B", "C", "D", "E"): 
            cell = ws[f'{col}{row_num}']
            cell.font = Font(bold=True)
            cell.border = self.double_rule

        for col in amount_columns:
            cell = ws[f'{col}{row_num}']
            cell.value = f'=SUM({col}{start_row}:{col}{end_row})'
            cell.font = Font(bold=True)
            cell.number_format ="#,##0.00_);(#,##0.00)"
            cell.border = self.double_rule

    def long_date(self, _date):
        if type(_date) == str:
            _date = date(int(_date[:4]), int(_date[5:7]), int(_date[-2:]))

        return _date.strftime("%B %d, %Y")


